import sqlite3

import openpyxl
from telebot.types import *

from core.token import bot
from core.filter import *
from core.xarid_parse import Xarid


def start(message):
    intro = """
Hello, I am `GRXarid` bot!
My tasks:
---
parse `https://api.xt-xarid.uz/urpc`
---
please press to >> /help << for more info!
    """
    if user(message) is False:
        db = sqlite3.connect("bot.db")
        db.execute(
            f"insert into User (ID, Username, Name, Last_Name) values ({message.chat.id}, '{message.chat.username}', '{message.chat.first_name}', '{message.chat.last_name}')"
        )
        db.commit()
        db.close()
    bot.send_message(message.chat.id, intro)


def usage(message):
    info = """
_commands_:
/start
/help
/product
_examples_:
`/product 60000 200`
---
second parameter > 200:
this should be more `10` and less `10000`
    """
    bot.send_message(message.chat.id, info)


def parse_id(message):
    id = int(message.text.split(" ")[1])
    create_excel_file(message)
    excel_row = 2
    bot.send_message(message.chat.id, f"Please wait {2*int(message.text.split()[2])}~ seconds...")
    data_return = False
    for num in range(id, id+int(message.text.split()[2])):
        data = Xarid(num).parse()
        if data is False:
            continue
        elif data == "Nope":
            continue
        else:
            print("[*] return data")
            data_return = True
            write_to_excel(message, data, excel_row)
            excel_row += 1
            bot.send_message(message.chat.id, f"ID:\t{data[0]}\nИмя товара:\t{data[1]}\nСтрана:\t{data[3]}\nСрок торга:\t{data[4]}\nЦена:\t{data[5]}\nКоличество:\t{data[6]}\nРайоны:\t{data[7]}\nСсылка:\t{data[8]}")
    if data_return is False:
        bot.send_message(message.chat.id, "All id was invalid types or not have info about id")
    else:
        file = open(f"./Files/{message.chat.id}_{message.date}.xlsx", "rb")
        document = file.read()
        file.close()
        bot.send_document(message.chat.id, document, visible_file_name=f"{message.chat.id}_{message.date}.xlsx")


def create_excel_file(message) -> None:
    wb = openpyxl.Workbook()
    sheet = wb.active
    __values = (
        "ID",
        "Имя товара",
        "Описание",
        "Страна производства",
        "Срок торга",
        "Цена",
        "Количество",
        "Районы",
        "Ссылка"
    )
    for num in range(0, len(__values)):
        sheet.cell(row=1, column=num+1).value = __values[num]
    wb.save(filename=f"./Files/{message.chat.id}_{message.date}.xlsx")


def write_to_excel(message, data, excel_row) -> None:
    wb = openpyxl.load_workbook(f"./Files/{message.chat.id}_{message.date}.xlsx")
    sheet = wb.active
    for num in range(len(data)):
        sheet.cell(row=excel_row, column=num+1).value = data[num]
    wb.save(filename=f"./Files/{message.chat.id}_{message.date}.xlsx")